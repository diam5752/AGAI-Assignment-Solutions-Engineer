"""Integration-style tests that exercise the CLI pipeline entrypoint."""
import csv
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from automation import pipeline
from automation.cli import main as cli_main


@pytest.fixture(autouse=True)
def _reset_argv(monkeypatch: pytest.MonkeyPatch) -> None:
    """Ensure sys.argv starts clean for each CLI invocation."""

    monkeypatch.setattr(sys, "argv", ["automation.cli"])


def _run_cli(monkeypatch: pytest.MonkeyPatch, args: list[str]) -> None:
    monkeypatch.setattr(sys, "argv", ["automation.cli", *args])
    cli_main()


def test_cli_writes_csv_output(tmp_path: Path, dummy_data_dir: Path, expected_record_count: int, monkeypatch: pytest.MonkeyPatch) -> None:
    csv_output = tmp_path / "records.csv"

    _run_cli(
        monkeypatch,
        ["--data-dir", str(dummy_data_dir), "--output", str(csv_output), "--sink", "csv"],
    )

    with csv_output.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    assert len(rows) == expected_record_count


def test_cli_writes_excel_output(tmp_path: Path, dummy_data_dir: Path, expected_record_count: int, monkeypatch: pytest.MonkeyPatch) -> None:
    csv_output = tmp_path / "records.csv"
    excel_output = tmp_path / "records.xlsx"

    _run_cli(
        monkeypatch,
        [
            "--data-dir",
            str(dummy_data_dir),
            "--output",
            str(csv_output),
            "--sink",
            "excel",
            "--excel-output",
            str(excel_output),
        ],
    )

    workbook = load_workbook(excel_output)
    sheet = workbook.active
    assert sheet.title == "unified_records"
    assert sheet.max_row - 1 == expected_record_count


def test_cli_sheets_sink_uses_pipeline(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, dummy_data_dir: Path, expected_record_count: int, fake_service_account_file: Path) -> None:
    csv_output = tmp_path / "records.csv"
    captured: dict[str, list[dict[str, str]] | None] = {"rows": None}

    def fake_push(rows, **_: str):
        captured["rows"] = list(rows)

    monkeypatch.setattr(pipeline, "push_to_google_sheets", fake_push)

    _run_cli(
        monkeypatch,
        [
            "--data-dir",
            str(dummy_data_dir),
            "--output",
            str(csv_output),
            "--sink",
            "sheets",
            "--spreadsheet-id",
            "dummy",
            "--service-account",
            str(fake_service_account_file),
        ],
    )

    assert captured["rows"] is not None
    assert len(captured["rows"] or []) == expected_record_count
